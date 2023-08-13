import logging

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pretty_utils.miscellaneous.time_and_date import unix_to_strtime, strtime_to_unix
from pretty_utils.type_functions.lists import split_list

from data import config
from data.models import Statuses
from utils.db_api.database import get_account, db
from utils.db_api.tables import Account
from utils.miscellaneous.read_spreadsheet import read_spreadsheet


class General:
    @staticmethod
    def import_accounts() -> None:
        accounts = read_spreadsheet(path=config.ACCOUNTS_FILE)
        if accounts:
            print('Importing accounts...')
            imported = []
            edited = []
            total = len(accounts)
            account_lists = split_list(s_list=accounts, n=1000)
            for account_list in account_lists:
                insert_it = []
                for account in account_list:
                    try:
                        login = account.get('login')
                        password = account.get('password')
                        if not all((login, password,)):
                            continue

                        login = str(login).lower()
                        status = account.get('status')
                        status = status if status else Statuses.New
                        rank = account.get('rank')
                        rank = int(rank) if rank else None
                        last_online = account.get('last_online')
                        last_online = strtime_to_unix(
                            strtime=last_online, format='%d.%m.%Y %H:%M:%S'
                        ) if last_online else None
                        if login:
                            account_instance = get_account(login=login)
                            if account_instance and account_instance.password != password:
                                account_instance.password = password
                                account_instance.status = Statuses.New
                                db.commit()
                                edited.append(account_instance)

                            elif not account_instance:
                                account_instance = Account(
                                    login=login, password=password, status=status, rank=rank,
                                    last_online=last_online
                                )
                                insert_it.append(account_instance)
                                imported.append(account_instance)

                    except BaseException as e:
                        logging.exception('General.import_accounts')
                        print(f'{config.RED}Failed to import an account: {e}{config.RESET_ALL}')

                db.insert(insert_it)

            print(
                f'Done! {config.LIGHTGREEN_EX}{len(imported)}/{total}{config.RESET_ALL} accounts were imported, '
                f'password have been changed at {config.LIGHTGREEN_EX}{len(edited)}/{total}{config.RESET_ALL}.'
            )

    @staticmethod
    def export_accounts() -> None:
        accounts = list(db.execute('SELECT * FROM accounts'))
        if accounts:
            spreadsheet = Workbook()
            sheet: Worksheet = spreadsheet['Sheet']
            for column, header in enumerate(['n'] + list(db.execute('SELECT * FROM accounts').keys())[1:]):
                sheet.cell(row=1, column=column + 1).value = header

            for row, account in enumerate(accounts):
                cell = sheet.cell(row=row + 2, column=1)
                cell.number_format = '0'
                cell.value = row + 1

                for column, value in enumerate(account[1:]):
                    cell = sheet.cell(row=row + 2, column=column + 2)
                    if value and column == 4:
                        value = unix_to_strtime(value)
                        cell.number_format = 'DD.MM.YYYY HH:MM'

                    cell.value = value

            spreadsheet.save(config.ACCOUNTS_FILE)
            print(
                f'\nDone! {config.LIGHTGREEN_EX}{len(accounts)}{config.RESET_ALL} accounts were exported to the '
                f'{config.LIGHTGREEN_EX}accounts.xlsx{config.RESET_ALL} file.'
            )

        else:
            print(f"\n{config.RED}You don't have any accounts added to the DB!{config.RESET_ALL}")
